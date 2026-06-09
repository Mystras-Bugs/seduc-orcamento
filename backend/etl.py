"""ETL: ingests budget spreadsheets into SQLite.

File 1 – Expedientes Consolidados (Reduções e Suplementações):
    TIPO | Expediente | Envio | Conclusão | ASSUNTO | UO | PROG | AÇÃO |
    GRUPO | ELEMENTO | FONTE | REDUÇÃO | SUPLEMENTAÇÃO | DESCRIÇÃO |
    Valor minutado | JAN..DEZ

    The first row contains stray totals (often noise) and is skipped.
    The second row is the header.

File 2 – Planejado Consolidado (Execução Orçamentária):
    Columns: COD_UO, NOME_UO, COD_UGE, NOME_UGE, COD_FUNCAO, NOME_FUNCAO,
    COD_SUB_FUNCAO, NOME_SUB_FUNCAO, COD_PROGRAMA, NOME_PROGRAMA, COD_ACAO,
    NOME_ACAO, COD_FONTE_DETALHADA, NOME_FONTE_DETALHADA, COD_GRUPO, NOME_GRUPO,
    COD_ELEMENTO, NOME_ELEMENTO, ID_ASSUNTO_PLANEJADO, ASSUNTO_PLANEJADO,
    CLASSIFICACAO, VINCULACAO, PODE_SER_FUNDEB, TABELA, VALOR, MES_01..MES_12

    Each budget line is repeated 5 times with TABELA values:
    10 - DOTACAO, 30 - EMPENHADO, 35 - A EMPENHAR, 40 - LIQUIDADO, 45 - A LIQUIDAR
    First row is the header row.
"""
from __future__ import annotations

import re
from datetime import datetime, date
from pathlib import Path
from typing import Iterable

import openpyxl

from .database import get_conn

MESES = [
    "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]


def _norm_text(value) -> str | None:
    if value is None:
        return None
    s = str(value).replace("\r", " ").replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s or None


GRUPOS_VALIDOS = {"31", "33", "44"}


def _norm_grupo(value) -> str | None:
    """Padroniza o grupo de despesa.

    A SEDUC opera apenas tres grupos validos: 31 (Folha de Pessoal),
    33 (Custeio) e 44 (Capital). Lancamentos com codigos invalidos
    (34, 35 etc.) sao erros humanos de digitacao e devem ser absorvidos
    no grupo conceitualmente equivalente: qualquer 3X que nao seja 31
    vira 33 (Custeio); qualquer 4X que nao seja 44 vira 44 (Capital).
    """
    if value is None or value == "":
        return None
    digits = re.sub(r"\D", "", str(value))
    if not digits:
        return None
    code = digits[:2] if len(digits) >= 2 else digits.zfill(2)
    if code in GRUPOS_VALIDOS:
        return code
    if code.startswith("3"):
        return "33"
    if code.startswith("4"):
        return "44"
    return code


def _norm_acao(value) -> str | None:
    """Normaliza o lancamento de Acao para 'NNNN - NOME'.

    O codigo numerico no inicio (4 digitos) e o identificador real
    da acao. O nome textual pode variar entre planilhas (pontuacao,
    acento, abreviacao). Aqui apenas garantimos zero-padding no
    codigo; a unificacao do nome canonico acontece em SQL ao final
    do import.
    """
    if value is None or value == "":
        return None
    raw = " ".join(str(value).replace("\n", " ").split()).strip()
    m = re.match(r"^\s*(\d{1,4})\s*[-,.–]?\s*(.*)$", raw)
    if not m:
        return raw
    code = m.group(1).zfill(4)
    name = m.group(2).strip(" -,.–")
    return f"{code} - {name}" if name else code


def _norm_codigo_padded(value, width: int) -> str | None:
    """Normalises a numeric-style code to a fixed width with zero-padding.

    Handles common spreadsheet quirks:
        - int values     -> '815'    -> '0815'
        - float values   -> '815.0'  -> '0815'
        - string '  815' -> '0815'
        - non-numeric    -> returned as-is (rare; kept for safety)
    """
    if value is None or value == "":
        return None
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    digits = re.sub(r"\D", "", s)
    if digits:
        return digits.zfill(width)
    return s


def _norm_uo(value) -> tuple[str | None, str | None]:
    """Split '08011 - SUBSECRETARIA PEDAGOGICA' into (code, name).

    Many rows have slight code variations (08011 vs 8011) and punctuation
    differences. We normalise the code to a 5-digit padded form when possible.
    """
    raw = _norm_text(value)
    if not raw:
        return None, None
    m = re.match(r"^\s*(\d{4,5})\s*[-,]?\s*(.+)?$", raw)
    if not m:
        return None, raw
    code = m.group(1).zfill(5)
    name = (m.group(2) or "").strip(" -,.")
    return code, name or None


def _to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        s = str(value).replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0


def _to_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _infer_status(envio: str | None, conclusao: str | None) -> str:
    if conclusao:
        return "Realizada"
    if envio:
        return "Em Andamento"
    return "Em Elaboracao"


def _iter_rows(ws) -> Iterable[tuple]:
    rows = ws.iter_rows(values_only=True)
    # Skip first two rows (noise + header).
    for _ in range(2):
        try:
            next(rows)
        except StopIteration:
            return
    for row in rows:
        if not row or all(c is None or c == "" for c in row[:5]):
            continue
        yield row


def import_xlsx(path: str | Path, replace: bool = True) -> dict:
    """Read the spreadsheet and persist all rows.

    Returns a small summary dict for the API response.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    inserted = 0
    primeira_data = None
    ultima_data = None
    total_reducao = 0.0
    total_suplementacao = 0.0

    with get_conn() as conn:
        if replace:
            conn.execute("DELETE FROM expedientes")
        cur = conn.cursor()
        for row in _iter_rows(ws):
            tipo = _norm_text(row[0])
            expediente = _norm_text(row[1])
            envio = _to_date(row[2])
            conclusao = _to_date(row[3])
            assunto = _norm_text(row[4])
            uo_codigo, uo_nome = _norm_uo(row[5])
            programa = _norm_codigo_padded(row[6], 4)
            acao = _norm_acao(row[7])
            grupo = _norm_grupo(row[8])
            elemento = _norm_text(row[9])
            fonte = _norm_codigo_padded(row[10], 9)
            reducao = _to_float(row[11])
            suplementacao = _to_float(row[12])
            descricao = _norm_text(row[13])
            valor_minutado = _to_float(row[14])
            meses_vals = [_to_float(row[15 + i]) for i in range(12)]

            if expediente is None and assunto is None:
                continue

            status = _infer_status(envio, conclusao)

            cur.execute(
                """
                INSERT INTO expedientes (
                    tipo, expediente, data_envio, data_conclusao, assunto,
                    uo_codigo, uo_nome, programa, acao, grupo, elemento, fonte,
                    reducao, suplementacao, descricao, valor_minutado,
                    janeiro, fevereiro, marco, abril, maio, junho, julho,
                    agosto, setembro, outubro, novembro, dezembro, status
                ) VALUES (?,?,?,?,?, ?,?,?,?,?,?,?, ?,?,?,?, ?,?,?,?,?,?,?,?,?,?,?,?, ?)
                """,
                (
                    tipo, expediente, envio, conclusao, assunto,
                    uo_codigo, uo_nome, programa, acao, grupo, elemento, fonte,
                    reducao, suplementacao, descricao, valor_minutado,
                    *meses_vals,
                    status,
                ),
            )
            inserted += 1
            total_reducao += reducao
            total_suplementacao += suplementacao
            for d in (envio, conclusao):
                if d:
                    primeira_data = d if not primeira_data or d < primeira_data else primeira_data
                    ultima_data = d if not ultima_data or d > ultima_data else ultima_data

        # Higieniza nomes de UO: o mesmo codigo pode aparecer com pequenas
        # variacoes (acentuacao, pontuacao, abreviacoes). Escolhemos o nome
        # canonico pela maior ocorrencia (com maior comprimento como
        # desempate) e aplicamos a todas as linhas com o mesmo codigo.
        conn.execute(
            """
            UPDATE expedientes
               SET uo_nome = (
                   SELECT uo_nome FROM expedientes e2
                    WHERE e2.uo_codigo = expedientes.uo_codigo
                      AND e2.uo_nome IS NOT NULL
                    GROUP BY uo_nome
                    ORDER BY COUNT(*) DESC, LENGTH(uo_nome) DESC, uo_nome
                    LIMIT 1
               )
             WHERE uo_codigo IS NOT NULL
            """
        )

        # Higieniza nomes de Acao: o codigo de 4 digitos no inicio define a
        # acao real; o nome textual pode variar entre lancamentos. Escolhemos
        # como canonico o nome com maior ocorrencia para cada codigo.
        conn.execute(
            """
            UPDATE expedientes
               SET acao = (
                   SELECT acao FROM expedientes e2
                    WHERE substr(e2.acao, 1, 4) = substr(expedientes.acao, 1, 4)
                      AND e2.acao IS NOT NULL
                    GROUP BY acao
                    ORDER BY COUNT(*) DESC, LENGTH(acao) DESC, acao
                    LIMIT 1
               )
             WHERE acao IS NOT NULL
            """
        )

        conn.execute(
            "INSERT INTO importacoes (arquivo, linhas, observacao) VALUES (?,?,?)",
            (path.name, inserted, f"replace={replace}"),
        )

    return {
        "arquivo": path.name,
        "linhas_importadas": inserted,
        "periodo": {"de": primeira_data, "ate": ultima_data},
        "total_reducao": round(total_reducao, 2),
        "total_suplementacao": round(total_suplementacao, 2),
    }


# ---------------------------------------------------------------------------
# ETL 2 – Planejado Consolidado (Execução Orçamentária)
# ---------------------------------------------------------------------------

# TABELA column codes that exist in the planejado file
_TABELA_MAP = {
    "10": "DOTACAO",
    "30": "EMPENHADO",
    "35": "A_EMPENHAR",
    "40": "LIQUIDADO",
    "45": "A_LIQUIDAR",
}


def _tabela_code(raw: str | None) -> str | None:
    """Extract the numeric prefix from 'NN - NAME' TABELA values."""
    if not raw:
        return None
    m = re.match(r"^\s*(\d+)", str(raw))
    return m.group(1).strip() if m else None


def import_execucao_xlsx(path: str | Path, replace: bool = True) -> dict:
    """Read the PLANEJADO_CONSOLIDADO spreadsheet and persist all rows.

    Each unique combination of (UO, UGE, Programa, Ação, Fonte, Grupo,
    Elemento, Assunto) appears 5 times with different TABELA values.
    Each such row is inserted individually so the frontend can pivot as needed.

    Returns a small summary dict for the API response.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    inserted = 0
    total_dotacao = 0.0
    total_empenhado = 0.0

    with get_conn() as conn:
        if replace:
            conn.execute("DELETE FROM execucao_orcamentaria")
        cur = conn.cursor()

        rows = ws.iter_rows(values_only=True)
        try:
            next(rows)  # skip header row
        except StopIteration:
            return {"arquivo": path.name, "linhas_importadas": 0}

        for row in rows:
            if not row or all(c is None or c == "" for c in row[:5]):
                continue

            uo_codigo              = _norm_codigo_padded(row[0], 5)
            uo_nome                = _norm_text(row[1])
            uge_codigo             = _norm_codigo_padded(row[2], 6)
            uge_nome               = _norm_text(row[3])
            funcao_codigo          = _norm_codigo_padded(row[4], 2)
            funcao_nome            = _norm_text(row[5])
            subfuncao_codigo       = _norm_codigo_padded(row[6], 3)
            subfuncao_nome         = _norm_text(row[7])
            programa               = _norm_codigo_padded(row[8], 4)
            programa_nome          = _norm_text(row[9])
            acao                   = _norm_codigo_padded(row[10], 4)
            acao_nome              = _norm_text(row[11])
            fonte                  = _norm_codigo_padded(row[12], 9)
            fonte_nome             = _norm_text(row[13])
            grupo                  = _norm_grupo(row[14])
            grupo_nome             = _norm_text(row[15])
            elemento               = _norm_text(row[16])
            elemento_nome          = _norm_text(row[17])
            id_assunto_planejado   = _norm_text(row[18])
            assunto                = _norm_text(row[19])
            classificacao          = _norm_text(row[20])
            vinculacao             = _norm_text(row[21])
            pode_ser_fundeb        = _norm_text(row[22])
            tabela                 = _norm_text(row[23])
            valor                  = _to_float(row[24])

            # Monthly distribution columns MES_01..MES_12 (index 25..36)
            meses_vals = [_to_float(row[25 + i]) for i in range(12)]

            if uo_codigo is None and assunto is None:
                continue

            # Track totals for the summary
            tc = _tabela_code(tabela)
            if tc == "10":
                total_dotacao += valor
            elif tc == "30":
                total_empenhado += valor

            cur.execute(
                """
                INSERT INTO execucao_orcamentaria (
                    uo_codigo, uo_nome, uge_codigo, uge_nome,
                    funcao_codigo, funcao_nome, subfuncao_codigo, subfuncao_nome,
                    programa, programa_nome, acao, acao_codigo, acao_nome,
                    fonte, fonte_nome, grupo, grupo_nome,
                    elemento, elemento_nome, id_assunto_planejado, assunto,
                    classificacao, vinculacao, pode_ser_fundeb,
                    tabela, valor,
                    mes_01, mes_02, mes_03, mes_04, mes_05, mes_06,
                    mes_07, mes_08, mes_09, mes_10, mes_11, mes_12
                ) VALUES (
                    ?,?,?,?, ?,?,?,?, ?,?,?,?,?, ?,?,?,?, ?,?,?,?, ?,?,?, ?,?,
                    ?,?,?,?,?,?, ?,?,?,?,?,?
                )
                """,
                (
                    uo_codigo, uo_nome, uge_codigo, uge_nome,
                    funcao_codigo, funcao_nome, subfuncao_codigo, subfuncao_nome,
                    programa, programa_nome, acao, acao, acao_nome,
                    fonte, fonte_nome, grupo, grupo_nome,
                    elemento, elemento_nome, id_assunto_planejado, assunto,
                    classificacao, vinculacao, pode_ser_fundeb,
                    tabela, valor,
                    *meses_vals,
                ),
            )
            inserted += 1

        conn.execute(
            "INSERT INTO importacoes (arquivo, linhas, observacao) VALUES (?,?,?)",
            (path.name, inserted, f"tipo=execucao replace={replace}"),
        )

    return {
        "arquivo": path.name,
        "linhas_importadas": inserted,
        "total_dotacao": round(total_dotacao, 2),
        "total_empenhado": round(total_empenhado, 2),
    }


# ---------------------------------------------------------------------------
# ETL 3 – Necessidade Atualizada (por UGE e Assunto Planejado)
# ---------------------------------------------------------------------------

def _norm_header(value) -> str:
    """Normaliza um cabeçalho para casar com os aliases do COLMAP:
    maiusculas, sem acento, espacos/underscores/pontos unificados em '_'."""
    if value is None:
        return ""
    s = str(value).strip().upper()
    acentos = (
        ("Á", "A"), ("À", "A"), ("Â", "A"), ("Ã", "A"), ("Ä", "A"),
        ("É", "E"), ("Ê", "E"), ("È", "E"), ("Ë", "E"),
        ("Í", "I"), ("Î", "I"), ("Ì", "I"), ("Ï", "I"),
        ("Ó", "O"), ("Ô", "O"), ("Õ", "O"), ("Ò", "O"), ("Ö", "O"),
        ("Ú", "U"), ("Û", "U"), ("Ù", "U"), ("Ü", "U"), ("Ç", "C"),
    )
    for a, b in acentos:
        s = s.replace(a, b)
    s = re.sub(r"[\s\.\-/]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def import_necessidade_xlsx(path: str | Path, replace: bool = True) -> dict:
    """Read the NECESSIDADE spreadsheet (necessidade atualizada por UGE e
    assunto planejado) and persist into the `necessidade` table.

    O casamento de colunas é feito por NOME de cabeçalho (não por posição),
    usando o COLMAP abaixo. Cada campo aceita vários aliases já normalizados
    por _norm_header — ajuste-os aqui caso a planilha use outros rótulos.
    """
    # campo_destino -> lista de aliases de cabeçalho aceitos (normalizados)
    COLMAP: dict[str, list[str]] = {
        "uge_codigo":             ["COD_UGE", "UGE", "CODIGO_UGE", "COD_UNIDADE_GESTORA", "UGE_CODIGO"],
        "uge_nome":               ["NOME_UGE", "UNIDADE_GESTORA", "NOME_UNIDADE_GESTORA", "UGE_NOME"],
        "uo_codigo":              ["COD_UO", "UO", "CODIGO_UO", "UO_CODIGO", "COD_UNIDADE_ORCAMENTARIA"],
        "uo_nome":                ["NOME_UO", "UNIDADE_ORCAMENTARIA", "UO_NOME"],
        "programa":               ["COD_PROGRAMA", "PROGRAMA", "PROG"],
        "acao":                   ["COD_ACAO", "ACAO", "ACAO_CODIGO"],
        "grupo":                  ["COD_GRUPO", "GRUPO", "GRUPO_DESPESA", "GRUPO_DE_DESPESA"],
        "elemento":               ["COD_ELEMENTO", "ELEMENTO"],
        "fonte":                  ["COD_FONTE_DETALHADA", "FONTE", "COD_FONTE", "FONTE_DETALHADA"],
        "id_assunto_planejado":   ["ID_ASSUNTO_PLANEJADO", "ID_ASSUNTO", "ID_ASSUNTO_PLAN"],
        "assunto":                ["ASSUNTO_PLANEJADO", "ASSUNTO", "DESPESA", "DESPESA_ASSUNTO"],
        "necessidade_atualizada": ["NECESSIDADE_ATUALIZADA", "NECESSIDADE", "VALOR_NECESSIDADE",
                                    "NECESSIDADE_ATUAL", "NECESSIDADE_TOTAL"],
        "observacao":             ["OBSERVACAO", "OBS", "OBSERVACOES", "OBSERVACAO_GERAL"],
    }

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {"arquivo": path.name, "linhas_importadas": 0,
                "colunas_nao_encontradas": list(COLMAP), "cabecalho": []}

    # Resolve cabeçalho -> índice de coluna por nome
    headers_norm = [_norm_header(h) for h in header]
    header_index: dict[str, int] = {}
    for i, h in enumerate(headers_norm):
        if h and h not in header_index:
            header_index[h] = i

    col_idx: dict[str, int] = {}
    nao_encontradas: list[str] = []
    for campo, aliases in COLMAP.items():
        for alias in aliases:
            if alias in header_index:
                col_idx[campo] = header_index[alias]
                break
        else:
            nao_encontradas.append(campo)

    def cell(row, campo):
        i = col_idx.get(campo)
        if i is None or i >= len(row):
            return None
        return row[i]

    inserted = 0
    total_necessidade = 0.0

    with get_conn() as conn:
        if replace:
            conn.execute("DELETE FROM necessidade")
        cur = conn.cursor()

        for row in rows:
            if not row or all(c is None or c == "" for c in row):
                continue

            uge_codigo = _norm_codigo_padded(cell(row, "uge_codigo"), 6)
            uge_nome   = _norm_text(cell(row, "uge_nome"))
            uo_codigo, uo_nome_split = _norm_uo(cell(row, "uo_codigo"))
            uo_nome    = _norm_text(cell(row, "uo_nome")) or uo_nome_split
            programa   = _norm_codigo_padded(cell(row, "programa"), 4)
            acao       = _norm_codigo_padded(cell(row, "acao"), 4)
            grupo      = _norm_grupo(cell(row, "grupo"))
            elemento   = _norm_text(cell(row, "elemento"))
            fonte      = _norm_codigo_padded(cell(row, "fonte"), 9)
            id_assunto = _norm_text(cell(row, "id_assunto_planejado"))
            assunto    = _norm_text(cell(row, "assunto"))
            necessidade = _to_float(cell(row, "necessidade_atualizada"))
            observacao = _norm_text(cell(row, "observacao"))

            if assunto is None and id_assunto is None and uge_codigo is None:
                continue

            cur.execute(
                """
                INSERT INTO necessidade (
                    uge_codigo, uge_nome, uo_codigo, uo_nome,
                    programa, acao, grupo, elemento, fonte,
                    id_assunto_planejado, assunto, necessidade_atualizada, observacao
                ) VALUES (?,?,?,?, ?,?,?,?,?, ?,?,?,?)
                """,
                (
                    uge_codigo, uge_nome, uo_codigo, uo_nome,
                    programa, acao, grupo, elemento, fonte,
                    id_assunto, assunto, necessidade, observacao,
                ),
            )
            inserted += 1
            total_necessidade += necessidade

        # Higieniza nomes de UO/UGE pelo nome canônico mais frequente por código
        conn.execute(
            """
            UPDATE necessidade
               SET uo_nome = (
                   SELECT uo_nome FROM necessidade e2
                    WHERE e2.uo_codigo = necessidade.uo_codigo
                      AND e2.uo_nome IS NOT NULL
                    GROUP BY uo_nome
                    ORDER BY COUNT(*) DESC, LENGTH(uo_nome) DESC, uo_nome
                    LIMIT 1)
             WHERE uo_codigo IS NOT NULL
            """
        )
        conn.execute(
            """
            UPDATE necessidade
               SET uge_nome = (
                   SELECT uge_nome FROM necessidade e2
                    WHERE e2.uge_codigo = necessidade.uge_codigo
                      AND e2.uge_nome IS NOT NULL
                    GROUP BY uge_nome
                    ORDER BY COUNT(*) DESC, LENGTH(uge_nome) DESC, uge_nome
                    LIMIT 1)
             WHERE uge_codigo IS NOT NULL
            """
        )

        conn.execute(
            "INSERT INTO importacoes (arquivo, linhas, observacao) VALUES (?,?,?)",
            (path.name, inserted, f"tipo=necessidade replace={replace}"),
        )

    return {
        "arquivo": path.name,
        "linhas_importadas": inserted,
        "total_necessidade": round(total_necessidade, 2),
        "colunas_mapeadas": {k: headers_norm[v] for k, v in col_idx.items()},
        "colunas_nao_encontradas": nao_encontradas,
        "cabecalho": [h for h in headers_norm if h],
    }
